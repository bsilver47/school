import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

running_path = "ECE3710_Quiz4_BenjaminSilver.xlsx"
running_object = openpyxl.load_workbook(running_path)
running_sheet_object = running_object.active

probA = 0
for i in range(2, 6):
    running_var = running_sheet_object.cell(row = i, column = 2).value
    if running_var <= 4:
        probA += 1
running_sheet_object.cell(row = 8, column = 2).value = probA

probB = 0
for i in range(2, 6):
    running_var = running_sheet_object.cell(row = i, column = 2).value
    if (1 <= running_var) and (running_var < 3):
        probB += 1
running_sheet_object.cell(row = 9, column = 2).value = probB

totalC = 0
for i in range(2, 6):
    running_var = running_sheet_object.cell(row = i, column = 1).value
    totalC += running_var
meanC = totalC / 5
running_sheet_object.cell(row = 10, column = 2).value = meanC

#please note that all values of x are currently sorted
#middle value of 5 = third val
medianD = running_sheet_object.cell(row = 4, column = 1).value
running_sheet_object.cell(row = 11, column = 2).value = medianD

totalE = 0
for i in range(2, 6):
    running_var = running_sheet_object.cell(row = i, column = 2).value
    totalE += running_var
meanE = totalE / 5
running_sheet_object.cell(row = 2, column = 4).value = meanC

stdE = "=STDEV(B2:B6)"
running_sheet_object.cell(row = 2, column = 5).value = ArrayFormula(stdE.strip('"'))

pmfE = "=NORMDIST(B2:B6,D2,E2,FALSE)"
running_sheet_object["B12"] = ArrayFormula(pmfE.strip('"'))
#running_sheet_object.cell(row = 12, column = 2).value = pmfE

cdfF = "=NORMDIST(B2:B6,D2,E2,TRUE)"
running_sheet_object["B17"] = ArrayFormula(cdfF.strip('"'))
#running_sheet_object.cell(row = 13, column = 2).value = cdfF

print("The update should have succeeded as of...")
running_object.save(running_path)
print("now.")
