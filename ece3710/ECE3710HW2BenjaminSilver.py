import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

running_path = "ECE3710HW2BenjaminSilver.xlsx"
running_object = openpyxl.load_workbook(running_path)
running_sheet_object = running_object.active

probOne = 0
for i in range(2, 3):
    running_var = running_sheet_object.cell(row = 2, column = i).value
    probOne += running_var
probNotOne = 1 - probOne
running_sheet_object.cell(row = 2, column = 5).value = probNotOne

probTwo = running_sheet_object.cell(row = 5, column = 2).value - running_sheet_object.cell(row = 5, column = 3).value
running_sheet_object.cell(row = 2, column = 4).value = probTwo

#p(a) = 3/10 because their are three defective components out of ten
probA = (3/10)
running_sheet_object.cell(row = 8, column = 2).value = probA
#p(b|a) = 2/9 because their are one less of each
probB = (2/9)
running_sheet_object.cell(row = 8, column = 6).value = probB
probAnB = probA * probB
running_sheet_object.cell(row = 8, column = 3).value = probAnB
probAPrimenB = (1 - probA) * probB
running_sheet_object.cell(row = 8, column = 5).value = probAPrimenB
running_sheet_object.cell(row = 8, column = 2).value = "The probability of B is dependent on A (as it is derived from the remaining possibilities after A has been taken into account), so the two are not independent of each other."

q4x = []
q4px = []
for i in range(3, 7):
    q4x.append(running_sheet_object.cell(row = 11, column = i).value)
    q4px.append(running_sheet_object.cell(row = 11, column = i).value)
q4a = 0
q4b = 0
for i in q4x:
    if i <= 2:
        q4a += 1
    if i > 1:
        q4b += 1
running_sheet_object.cell(row = 13, column = 3).value = q4a / len(q4x)
running_sheet_object.cell(row = 14, column = 3).value = q4b / len(q4x)
q4c = 0
for i in q4px:
    q4c += i
running_sheet_object.cell(row = 15, column = 3).value = q4c / len(q4px)

c = 1 / (1+2+3+4+5)
def q5p(x):
    if x in [1, 2, 3, 4, 5]:
        retval = c * x
    else:
        retval = 0
    return retval
running_sheet_object.cell(row = 17, column = 3).value = c
running_sheet_object.cell(row = 18, column = 3).value = q5p(2)
q4mean = 0
q4variance = 0
for i in range(1, 5):
    q4mean += q5p(i)
    q4variance += (q5p(i) ** 2)
q4mean = q4mean / 5
q4variance = q4variance / 5
running_sheet_object.cell(row = 19, column = 3).value = q4mean
running_sheet_object.cell(row = 20, column = 3).value = q4variance
running_sheet_object.cell(row = 21, column = 3).value = 4 # It kind of just feels like this is the right answer (assuming I have understood what is being asked; obviously there was more thought invested in this answer than that, it would take a while to explain how I got there though)
running_sheet_object.cell(row = 22, column = 3).value = 4 # If I understood the setup for this question, this is more or less what I thought was being asked for cdf. If my answers before e were correct, though, the answer to this problem would be 3
running_sheet_object.cell(row = 23, column = 3).value = 4 # Similar rationale to e

c = 1
def q6p(x):
    if (x > 0) and (x < 10):
        retval = c * x
    else:
        retval = 0
    return retval
running_sheet_object.cell(row = 25, column = 3).value = c
q6b = 0
for i in range(3, 9):
    q6b += q6p(i)
running_sheet_object.cell(row = 26, column = 3).value = q6b
q6mean = 0
for i in range(1, 9):
    q6mean += q6p(i)
q6mean = q6mean / 9
running_sheet_object.cell(row = 27, column = 3).value = q6mean
running_sheet_object.cell(row = 28, column = 3).value = 5
running_sheet_object.cell(row = 29, column = 3).value = 6

q7mx = 9.5
q7my = 6.8
q7sdx = 0.4
q7sdy = 0.1
running_sheet_object.cell(row = 31, column = 3).value = 3 * q7mx
running_sheet_object.cell(row = 31, column = 4).value = 3 * q7sdx
running_sheet_object.cell(row = 32, column = 3).value = q7my - q7mx
running_sheet_object.cell(row = 32, column = 4).value = q7sdy - q7sdx
running_sheet_object.cell(row = 33, column = 3).value = q7mx + (4 * q7my)
running_sheet_object.cell(row = 33, column = 4).value = q7sdx + (4 * q7sdy)


print("The update should have succeeded as of...")
running_object.save(running_path)
print("now.")
