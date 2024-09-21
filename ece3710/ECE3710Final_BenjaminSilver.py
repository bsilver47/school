# Benjamin Silver
# 10718564
# ECE3710 Final Exam

import scipy.stats as st
import numpy as np
import openpyxl

# 1.
a1Mu = 101.5
a1H0 = 100
a1n = 100
a1stdv = 25
a1z = (a1Mu - a1H0) / (a1stdv * (a1n ** 0.5))
pa1 = (1 - st.norm.cdf(a1z)) * 2
print("1. a)", pa1)
print("As the Probability is much higher than 0.05, we do not have sufficient data to disregard this model.")

pb1 = 1 - st.norm.cdf(a1z)
print("1. b)", pb1)
print("As the Probability is much higher than 0.05, we do not have sufficient data to disregard this model.")

c1n = 10
c1t = (a1Mu - a1H0) / (a1stdv * (c1n ** 0.5))
pc1 = (1 - st.norm.cdf(c1t)) * 2
print("1. c)", pc1)
print("As the Probability is much higher than 0.05, we do not have sufficient data to disregard this model.")


# 2.
x2 = [1, 2, 3]
y2 = [2, 3, 1]
n2 = 3
ave2x = (1 + 2 + 3) / n2
ave2y = (2 + 3 + 1) / n2
sum2 = 0
for i in range(n2):
    sum2 += ((x2[i] - ave2y) ** 2)
stdv2 = ((sum2 / n2) ** 0.5)
covar2 = stdv2 * stdv2 / n2
cro2 = (covar2/stdv2)*(stdv2/covar2)
print("2.", cro2)


# 3
in1x3 = [3, 4, 5, 6, 7]
in2x3 = [-1, 5, 11, 10, 7]
y3 = np.array([9, 7, 8, 7.6, 12])

def func3(i, a, b, c, d=0, e=0):
    function3 = (c * in2x3[i]) + (a * in1x3[i] * in2x3[i]) + (b * (in1x3[i] ** 2)) + d + e
    return function3

yhat3 = []
for i in range(5):
    temp3 = []
    for j in range(5):
        temp3.append(in2x3[i])
        temp3.append(in1x3[i] * in2x3[i])
        temp3.append((in1x3[i] ** 2))
    yhat3.append(temp3)
npyhat3 = np.array(yhat3)
npletter3 = np.array(["a", "b", "c", "d"])
npe3 = np.array(["e1", "e2", "e3", "e4", "e5"])
print("3. a)")
print(y3, "=", npyhat3, npletter3, "+", npe3)

print("3. b)")
running_path = "ECE3710Final.xlsx"
running_object = openpyxl.load_workbook(running_path)
running_sheet_object = running_object.active
probOne = 0
for i in range(1, 5):
    running_sheet_object.cell(row = i, column = 1).value = in1x3[i]
    running_sheet_object.cell(row = i, column = 2).value = in2x3[i]
    running_sheet_object.cell(row = i, column = 3).value = y3[i]
running_sheet_object.cell(row = 8, column = 3).value = "=LINEST(C1:C5,B1:B5, 1,0)"
print("running_sheet_object.cell(row = 8, column = 3).value = '=LINEST(C1:C5,B1:B5, 1,0)'")
print("3. c) y = 9.074 - 0.0553x")
d3 = '''
9 = -c - 3a + 9b +d
7 = 5c + 20a + 16b + d
8 = 11c + 55a + 25b + d
7.6 = 10c + 60a + 36b + d
12 = 7c + 49a + 49b + d
'''
print("3. d)", d3)

# 4
n4 = 50
ave4 = 9.5
stdv4 = 5
c55q4 = 55
c55a4 = 45
c55a2 = 22.5
z55 = st.norm.ppf(1-c55a2)
conf55plus = ave4 + (z55 * stdv4 / (n4 ** 0.5))
conf55minus = ave4 - (z55 * stdv4 / (n4 ** 0.5))
conf55 = [conf55minus, conf55plus]
c85q4 = 85
c85a4 = 15
c85a2 = 7.5
z85 = st.norm.ppf(1-c85a2)
conf85plus = ave4 + (z85 * stdv4 / (n4 ** 0.5))
conf85minus = ave4 - (z85 * stdv4 / (n4 ** 0.5))
conf85 = [conf85minus, conf85plus]
print("4.")
print("The mean value of a 55% confidence level is", conf55)
print("The mean value of a 85% confidence level is", conf85)


# 5
print("5.")
x5 = np.array([0.43, 0.67, 0.40, 0.45, 0.80])
y5 = np.array([772, 735, 774, 769, 723])
print("a)", y5, "=", x5, "x + b + e")
av5x = np.mean(x5)
av5y = np.mean(y5)
dev5xy = np.sum(y5)*np.sum(x5) - 5 * av5y * av5x
dev5xx = np.sum(x5)*np.sum(x5) - 5 * av5x * av5x
a5 = dev5xy / dev5xx
b5 = av5y - a5 * av5x
print("b) y =", a5, "x +", b5)
print('''
c)
589.96
919.24
548.8
617.4
1097.6
''')
sum5 = 0
for i in range(5):
    sum5 += ((y5[i] - av5y) ** 2)
stdv5e = ((sum5 / 5) ** 0.5)
print("5. d) e =", stdv5e)
